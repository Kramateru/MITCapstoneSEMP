"""
Trainer Portal Routes
Handles batch management, course assignment, interaction reviewing, and coaching
"""

import csv
import io
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from fastapi import (
    APIRouter,
    Depends as FastAPIDepends,
    File,
    Form,
    HTTPException,
    Query,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
)
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload

from .. import auth_utils
from ..database import SessionLocal, get_db
from ..default_credentials import DEFAULT_TRAINEE_PASSWORD
from ..models import (
    batch_user_association,
    Batch,
    CertificateRecord,
    Course,
    CourseAssignment,
    Feedback,
    FeedbackType,
    MicrolearningAssignment,
    MicrolearningModule,
    MicrolearningUploadedAsset,
    MicrolearningTopicCategory,
    PerformanceMetrics,
    PracticeSession,
    Scenario,
    ScenarioDifficulty,
    User,
    UserRole,
)
from ..services.certificate_awards import sync_trainee_completion_certificates
from ..services.live_updates import live_update_manager
from ..services.microlearning import (
    assignment_is_current,
    ensure_module_exercises,
    filter_current_assignments,
    get_module_media_state,
    refresh_assignment_progress,
    serialize_microlearning_module,
    serialize_assignment_summary,
)
from ..services.microlearning_delete import (
    MicrolearningDeleteError,
    delete_microlearning_module_and_dependencies,
)
from ..services.microlearning_catalog import (
    build_type_specific_exercises,
    normalize_module_type,
    serialize_topic_category,
)
from ..services.supabase_auth_service import SupabaseUserSyncError, sync_user_to_supabase_auth
from ..supabase_client import get_supabase_client

router = APIRouter(prefix="/api/trainer", tags=["trainer"])
TRAINER_BULK_UPLOAD_TEMPLATE = "trainer-trainee-bulk-upload-template.xlsx"
SUPABASE_PUBLIC_OBJECT_MARKER = "/storage/v1/object/public/"


def Depends(dependency=None):
    """Default empty Depends() to DB session in this module."""
    return FastAPIDepends(get_db if dependency is None else dependency)


def _normalize_email(email: str) -> str:
    return email.strip().lower()


def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (header or "").strip().lower()).strip("_")


def _normalize_text_value(value: Any) -> str:
    return str(value or "").strip()


def _sync_trainee_to_supabase_or_raise(db: Session, trainee: User) -> None:
    try:
        sync_user_to_supabase_auth(db, trainee)
    except SupabaseUserSyncError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc


def _derive_batch_name(name: Optional[str], wave_number: Optional[int]) -> str:
    normalized_name = _normalize_text_value(name)
    if normalized_name:
        return normalized_name
    if wave_number is not None:
        return f"Batch {wave_number}"
    raise HTTPException(status_code=400, detail="Provide a batch name or batch number")


def _get_trainer_batch(
    db: Session,
    *,
    trainer_id: str,
    batch_id: str,
) -> Batch:
    batch = (
        db.query(Batch)
        .filter(Batch.id == batch_id, Batch.created_by == trainer_id)
        .first()
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    return batch


def _get_trainer_course(
    db: Session,
    *,
    trainer_id: str,
    course_id: str,
) -> Course:
    course = (
        db.query(Course)
        .filter(Course.id == course_id, Course.created_by == trainer_id)
        .first()
    )
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    return course


def _get_trainer_microlearning_module(
    db: Session,
    *,
    trainer_id: str,
    module_id: str,
) -> MicrolearningModule:
    module = (
        db.query(MicrolearningModule)
        .options(
            selectinload(MicrolearningModule.assessment_method),
            selectinload(MicrolearningModule.topic_category),
        )
        .filter(
            MicrolearningModule.id == module_id,
            MicrolearningModule.created_by == trainer_id,
            MicrolearningModule.is_active == True,
        )
        .first()
    )
    if not module:
        raise HTTPException(status_code=404, detail="Microlearning activity not found")
    return module


def _get_trainer_microlearning_topic_category(
    db: Session,
    *,
    trainer_id: str,
    category_id: str,
) -> MicrolearningTopicCategory:
    category = (
        db.query(MicrolearningTopicCategory)
        .filter(
            MicrolearningTopicCategory.id == category_id,
            MicrolearningTopicCategory.created_by == trainer_id,
            MicrolearningTopicCategory.is_active == True,
        )
        .first()
    )
    if not category:
        raise HTTPException(status_code=404, detail="Microlearning topic category not found")
    return category


def _get_microlearning_assignment_counts(
    db: Session,
    *,
    module_ids: List[str],
) -> Dict[str, int]:
    normalized_ids = [module_id for module_id in dict.fromkeys(module_ids) if module_id]
    if not normalized_ids:
        return {}

    assignments = (
        db.query(MicrolearningAssignment)
        .options(
            selectinload(MicrolearningAssignment.module),
            selectinload(MicrolearningAssignment.batch),
            selectinload(MicrolearningAssignment.trainee).selectinload(User.batches),
        )
        .filter(MicrolearningAssignment.module_id.in_(normalized_ids))
        .all()
    )

    counts: Dict[str, int] = {}
    for assignment in filter_current_assignments(assignments):
        counts[assignment.module_id] = counts.get(assignment.module_id, 0) + 1

    return counts


def _get_trainer_trainee(
    db: Session,
    *,
    trainer_id: str,
    trainee_id: str,
) -> User:
    trainee = (
        db.query(User)
        .join(batch_user_association, batch_user_association.c.user_id == User.id)
        .join(Batch, Batch.id == batch_user_association.c.batch_id)
        .filter(
            User.id == trainee_id,
            User.role == UserRole.TRAINEE,
            User.is_active.is_(True),
            Batch.created_by == trainer_id,
        )
        .first()
    )
    if not trainee:
        raise HTTPException(status_code=404, detail="Trainee not found")
    return trainee


def _get_trainer_session(
    db: Session,
    *,
    trainer_id: str,
    session_id: str,
) -> PracticeSession:
    session = db.query(PracticeSession).filter(PracticeSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    _get_trainer_trainee(db, trainer_id=trainer_id, trainee_id=session.user_id)
    return session


def _resolve_batch_lookup(
    db: Session,
    *,
    trainer_id: str,
    batch_lookup: str,
) -> Optional[Batch]:
    raw_lookup = _normalize_text_value(batch_lookup)
    if not raw_lookup:
        return None

    base_query = db.query(Batch).filter(Batch.created_by == trainer_id, Batch.is_active == True)
    direct_match = (
        base_query.filter(func.lower(Batch.name) == raw_lookup.lower()).first()
    )
    if direct_match:
        return direct_match

    numeric_match = re.search(r"(\d+)", raw_lookup)
    if numeric_match:
        wave_number = int(numeric_match.group(1))
        return (
            base_query
            .filter(Batch.wave_number == wave_number)
            .order_by(Batch.created_at.desc())
            .first()
        )

    return None


def _build_batch_department_label(batch: Batch) -> str:
    if batch.name:
        return batch.name
    if batch.wave_number is not None:
        return f"Wave {batch.wave_number}"
    return "Unassigned Batch"


def _sort_batches_for_display(batches: List[Batch]) -> List[Batch]:
    return sorted(
        batches,
        key=lambda batch: (
            batch.wave_number is None,
            batch.wave_number if batch.wave_number is not None else 0,
            (batch.name or "").lower(),
        ),
    )


def _apply_batch_profile(user: User, batch: Optional[Batch]) -> None:
    user.lob = batch.lob if batch else None
    user.department = _build_batch_department_label(batch) if batch else None


def _refresh_trainee_profile_from_batches(user: User) -> None:
    next_batch = _sort_batches_for_display(list(user.batches or []))
    _apply_batch_profile(user, next_batch[0] if next_batch else None)


def _serialize_batch(batch: Optional[Batch]) -> Optional[Dict[str, Any]]:
    if not batch:
        return None
    return {
        "id": batch.id,
        "name": batch.name,
        "wave_number": batch.wave_number,
        "lob": batch.lob,
        "start_date": batch.start_date,
        "end_date": batch.end_date,
    }


def _serialize_trainer_batch_summary(batch: Batch) -> Dict[str, Any]:
    return {
        "id": batch.id,
        "name": batch.name,
        "description": batch.description,
        "wave_number": batch.wave_number,
        "lob": batch.lob,
        "start_date": batch.start_date,
        "end_date": batch.end_date,
        "is_active": batch.is_active,
        "users_count": len([user for user in batch.users if user.role == UserRole.TRAINEE]),
        "created_at": batch.created_at,
    }


def _serialize_trainee(
    user: User,
    batch: Optional[Batch] = None,
    *,
    batches: Optional[List[Batch]] = None,
) -> Dict[str, Any]:
    trainee_batches = _sort_batches_for_display(
        batches if batches is not None else ([batch] if batch else list(user.batches or []))
    )
    primary_batch = batch or (trainee_batches[0] if trainee_batches else None)
    return {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
        "is_active": user.is_active,
        "lob": user.lob,
        "department": user.department,
        "batch": _serialize_batch(primary_batch),
        "batches": [_serialize_batch(existing_batch) for existing_batch in trainee_batches],
        "batch_ids": [existing_batch.id for existing_batch in trainee_batches],
        "batch_names": [
            _build_batch_department_label(existing_batch) for existing_batch in trainee_batches
        ],
    }


def _sanitize_asset_name(filename: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "-", (filename or "").strip())
    return cleaned.strip("-") or "asset.bin"


def _resolve_supabase_public_storage_target(value: Any) -> tuple[Optional[str], Optional[str]]:
    normalized_value = _normalize_text_value(value)
    if not normalized_value or SUPABASE_PUBLIC_OBJECT_MARKER not in normalized_value:
        return None, None

    suffix = normalized_value.split(SUPABASE_PUBLIC_OBJECT_MARKER, 1)[1].strip().lstrip("/")
    if "/" not in suffix:
        return None, None

    bucket_name, storage_path = suffix.split("/", 1)
    normalized_bucket = _normalize_text_value(bucket_name)
    normalized_path = _normalize_text_value(storage_path).lstrip("/")
    if not normalized_bucket or not normalized_path:
        return None, None

    return normalized_bucket, normalized_path


def _normalize_supabase_asset_reference(content_data: dict[str, Any]) -> tuple[Optional[str], Optional[str]]:
    stored_bucket = _normalize_text_value(content_data.get("asset_bucket"))
    stored_path = _normalize_text_value(content_data.get("asset_storage_path")).lstrip("/")
    inferred_bucket = None
    inferred_path = None

    for candidate in (
        content_data.get("asset_url"),
        content_data.get("audio_url"),
    ):
        inferred_bucket, inferred_path = _resolve_supabase_public_storage_target(candidate)
        if inferred_path:
            break

    if inferred_path and (
        not stored_path
        or stored_path != inferred_path
        or not stored_path.startswith("microlearning/")
    ):
        return inferred_bucket or stored_bucket or None, inferred_path

    return stored_bucket or inferred_bucket or None, stored_path or inferred_path or None


def _resolve_microlearning_asset_folder(
    *,
    module_type: Optional[str],
    content_type: Optional[str],
) -> str:
    normalized_type = normalize_module_type(module_type) if module_type else ""
    normalized_content_type = _normalize_text_value(content_type).lower()

    if normalized_type == "video" or normalized_content_type.startswith("video/"):
        return "videos"
    if normalized_type == "infographic" or normalized_content_type.startswith("image/"):
        return "images"
    if normalized_type in {"audio", "case_study"} or normalized_content_type.startswith("audio/"):
        return "audio"
    return "assets"


def _upload_microlearning_asset(
    *,
    db: Session,
    trainer_id: str,
    file_bytes: bytes,
    filename: str,
    content_type: Optional[str],
    module_id: Optional[str] = None,
    module_type: Optional[str] = None,
) -> Dict[str, Any]:
    sanitized = _sanitize_asset_name(filename)
    normalized_module_id = _normalize_text_value(module_id)
    normalized_module_type = normalize_module_type(module_type) if module_type else ""
    module_storage_segment = (
        normalized_module_id
        or f"draft-{normalized_module_type or 'asset'}"
    )
    storage_folder = _resolve_microlearning_asset_folder(
        module_type=normalized_module_type,
        content_type=content_type,
    )

    supabase_client = get_supabase_client()
    if not supabase_client.is_available:
        raise HTTPException(
            status_code=503,
            detail="Supabase storage is required for trainer microlearning video and media uploads.",
        )

    bucket_name = supabase_client.microlearning_bucket_name
    storage_path = f"microlearning/{storage_folder}/{trainer_id}/{module_storage_segment}/{sanitized}"
    asset_url = supabase_client.upload_microlearning_binary(
        module_id=module_storage_segment,
        trainer_id=trainer_id,
        filename=sanitized,
        file_data=file_bytes,
        content_type=content_type or "application/octet-stream",
        folder=storage_folder,
    )
    if not asset_url:
        raise HTTPException(
            status_code=503,
            detail="Supabase storage could not save the uploaded microlearning media asset.",
        )

    asset = MicrolearningUploadedAsset(
        trainer_id=trainer_id,
        filename=sanitized,
        content_type=content_type or "application/octet-stream",
        byte_size=len(file_bytes),
        file_bytes=file_bytes,
    )
    db.add(asset)
    db.commit()
    db.refresh(asset)

    return {
        "asset_url": asset_url,
        "asset_record_id": asset.id,
        "storage_backend": "supabase_storage",
        "storage_path": storage_path,
        "bucket_name": bucket_name,
        "signed_url_required": True,
        "content_type": content_type or "application/octet-stream",
        "byte_size": asset.byte_size,
    }


def _prepare_replaced_module_asset_cleanup(
    *,
    previous_content_data: dict[str, Any],
    next_content_data: dict[str, Any],
) -> tuple[Optional[tuple[str, str]], Optional[str]]:
    previous_bucket_name, previous_storage_path = _normalize_supabase_asset_reference(previous_content_data)
    next_bucket_name, next_storage_path = _normalize_supabase_asset_reference(next_content_data)
    previous_bucket_name = previous_bucket_name or get_supabase_client().microlearning_bucket_name
    next_bucket_name = next_bucket_name or get_supabase_client().microlearning_bucket_name
    cleanup_storage_target = None
    if previous_storage_path and (
        previous_storage_path != next_storage_path
        or previous_bucket_name != next_bucket_name
    ):
        cleanup_storage_target = (previous_bucket_name, previous_storage_path)

    previous_asset_record_id = _normalize_text_value(previous_content_data.get("asset_record_id"))
    next_asset_record_id = _normalize_text_value(next_content_data.get("asset_record_id"))
    cleanup_asset_record_id = (
        previous_asset_record_id
        if previous_asset_record_id and previous_asset_record_id != next_asset_record_id
        else None
    )

    return cleanup_storage_target, cleanup_asset_record_id


def _feedback_category_db_label(category: FeedbackType | str | None) -> str:
    if isinstance(category, FeedbackType):
        return category.name
    normalized = str(category or "").strip()
    return normalized.upper() if normalized else FeedbackType.PRONUNCIATION.name


def _feedback_category_api_value(category: FeedbackType | str | None) -> str:
    if isinstance(category, FeedbackType):
        return category.value
    normalized = str(category or "").strip().lower()
    return normalized or FeedbackType.PRONUNCIATION.value


def _build_microlearning_report_overview(
    db: Session,
    *,
    trainer_id: str,
) -> dict[str, Any]:
    modules = (
        db.query(MicrolearningModule)
        .options(selectinload(MicrolearningModule.topic_category))
        .filter(
            MicrolearningModule.created_by == trainer_id,
            MicrolearningModule.is_active == True,
        )
        .all()
    )
    assignments = (
        db.query(MicrolearningAssignment)
        .options(
            selectinload(MicrolearningAssignment.module).selectinload(MicrolearningModule.topic_category),
            selectinload(MicrolearningAssignment.trainee).selectinload(User.batches),
            selectinload(MicrolearningAssignment.batch),
            selectinload(MicrolearningAssignment.certificate),
        )
        .filter(MicrolearningAssignment.assigned_by == trainer_id)
        .order_by(MicrolearningAssignment.assigned_at.desc())
        .all()
    )
    assignments = filter_current_assignments(assignments)

    did_update = False
    serialized_assignments: list[dict[str, Any]] = []
    for assignment in assignments:
        did_update = ensure_module_exercises(assignment.module) or did_update
        before = (
            assignment.status,
            assignment.completion_percentage,
            assignment.completed_exercises,
            assignment.completed_at,
        )
        refresh_assignment_progress(assignment)
        after = (
            assignment.status,
            assignment.completion_percentage,
            assignment.completed_exercises,
            assignment.completed_at,
        )
        did_update = before != after or did_update
        serialized_assignments.append(serialize_assignment_summary(assignment))

    if did_update:
        db.commit()

    batch_progress: dict[str, dict[str, Any]] = {}
    trainee_progress: dict[str, dict[str, Any]] = {}
    recent_certificates: list[dict[str, Any]] = []
    score_values: list[float] = []
    certified_count = 0
    completed_count = 0
    passed_count = 0

    for row in serialized_assignments:
        score = float(row.get("average_score") or 0.0)
        has_progress = int(row.get("completed_exercises") or 0) > 0
        if has_progress:
            score_values.append(score)

        if row.get("status") in {"completed", "certified"}:
            completed_count += 1
        if row.get("is_passed"):
            passed_count += 1
        if row.get("certificate_id"):
            certified_count += 1
            recent_certificates.append(
                {
                    "assignment_id": row["id"],
                    "certificate_id": row.get("certificate_id"),
                    "certificate_no": row.get("certificate_no"),
                    "module_title": row.get("module_title") or row.get("title"),
                    "trainee_name": row.get("trainee_name"),
                    "issued_at": row.get("certificate_issued_at") or row.get("completed_at"),
                }
            )

        batch_key = row.get("batch_id") or f"user::{row.get('user_id')}"
        batch_bucket = batch_progress.setdefault(
            batch_key,
            {
                "batch_id": row.get("batch_id"),
                "batch_name": row.get("batch_name") or "Individual assignment",
                "batch_label": row.get("batch_label") or row.get("batch_name") or "Individual assignment",
                "trainee_count": set(),
                "assignment_count": 0,
                "completed_count": 0,
                "certified_count": 0,
                "scores": [],
            },
        )
        batch_bucket["assignment_count"] += 1
        batch_bucket["trainee_count"].add(row.get("user_id"))
        if has_progress:
            batch_bucket["scores"].append(score)
        if row.get("status") in {"completed", "certified"}:
            batch_bucket["completed_count"] += 1
        if row.get("certificate_id"):
            batch_bucket["certified_count"] += 1

        trainee_key = row.get("user_id") or row["id"]
        trainee_bucket = trainee_progress.setdefault(
            trainee_key,
            {
                "trainee_id": row.get("user_id"),
                "trainee_name": row.get("trainee_name") or "Unknown trainee",
                "batch_label": row.get("batch_label") or row.get("batch_name") or "Individual assignment",
                "assignment_count": 0,
                "completed_count": 0,
                "certified_count": 0,
                "scores": [],
                "latest_completed_at": row.get("completed_at"),
            },
        )
        trainee_bucket["assignment_count"] += 1
        if has_progress:
            trainee_bucket["scores"].append(score)
        if row.get("status") in {"completed", "certified"}:
            trainee_bucket["completed_count"] += 1
        if row.get("certificate_id"):
            trainee_bucket["certified_count"] += 1
        if row.get("completed_at"):
            existing_completed_at = trainee_bucket.get("latest_completed_at")
            if not existing_completed_at or str(row["completed_at"]) > str(existing_completed_at):
                trainee_bucket["latest_completed_at"] = row["completed_at"]

    recent_certificates.sort(
        key=lambda entry: str(entry.get("issued_at") or ""),
        reverse=True,
    )

    def _average(values: list[float]) -> float:
        if not values:
            return 0.0
        return round(sum(values) / len(values), 2)

    batch_rows = [
        {
            "batch_id": bucket["batch_id"],
            "batch_name": bucket["batch_name"],
            "batch_label": bucket["batch_label"],
            "trainee_count": len(bucket["trainee_count"]),
            "assignment_count": bucket["assignment_count"],
            "completed_count": bucket["completed_count"],
            "certified_count": bucket["certified_count"],
            "average_score": _average(bucket["scores"]),
            "pass_rate": round(
                (bucket["certified_count"] / bucket["assignment_count"] * 100)
                if bucket["assignment_count"]
                else 0.0,
                2,
            ),
        }
        for bucket in batch_progress.values()
    ]
    batch_rows.sort(key=lambda row: (row["pass_rate"], row["average_score"], row["assignment_count"]), reverse=True)

    trainee_rows = [
        {
            "trainee_id": bucket["trainee_id"],
            "trainee_name": bucket["trainee_name"],
            "batch_label": bucket["batch_label"],
            "assignment_count": bucket["assignment_count"],
            "completed_count": bucket["completed_count"],
            "certified_count": bucket["certified_count"],
            "average_score": _average(bucket["scores"]),
            "pass_rate": round(
                (bucket["certified_count"] / bucket["assignment_count"] * 100)
                if bucket["assignment_count"]
                else 0.0,
                2,
            ),
            "latest_completed_at": bucket["latest_completed_at"],
        }
        for bucket in trainee_progress.values()
    ]
    trainee_rows.sort(key=lambda row: (row["pass_rate"], row["average_score"], row["assignment_count"]), reverse=True)

    category_rows = []
    category_buckets: dict[str, dict[str, Any]] = {}
    for module in modules:
        category_key = getattr(module, "topic_category_id", None) or "uncategorized"
        bucket = category_buckets.setdefault(
            category_key,
            {
                "topic_category_id": getattr(module, "topic_category_id", None),
                "topic_category_name": getattr(getattr(module, "topic_category", None), "name", None) or "Uncategorized",
                "module_count": 0,
            },
        )
        bucket["module_count"] += 1
    category_rows = sorted(category_buckets.values(), key=lambda row: (row["module_count"], row["topic_category_name"]), reverse=True)

    return {
        "summary": {
            "topic_category_count": len(category_rows),
            "module_count": len(modules),
            "assignment_count": len(serialized_assignments),
            "completed_count": completed_count,
            "certified_count": certified_count,
            "average_score": _average(score_values),
            "pass_rate": round((passed_count / len(serialized_assignments) * 100) if serialized_assignments else 0.0, 2),
        },
        "topic_categories": category_rows,
        "batch_progress": batch_rows,
        "trainee_progress": trainee_rows,
        "recent_certificates": recent_certificates[:8],
        "assignments": serialized_assignments,
    }


def _create_trainee_account(
    db: Session,
    *,
    email: str,
    full_name: str,
    batch: Batch,
    password: Optional[str] = None,
) -> User:
    normalized_email = _normalize_email(email)
    if not normalized_email:
        raise ValueError("Email address is required")

    normalized_name = _normalize_text_value(full_name)
    if not normalized_name:
        raise ValueError("Full name is required")

    existing_user = db.query(User).filter(func.lower(User.email) == normalized_email).first()
    if existing_user:
        raise ValueError(f"Email {normalized_email} already exists")

    temporary_password = password or DEFAULT_TRAINEE_PASSWORD
    new_user = User(
        email=normalized_email,
        full_name=normalized_name,
        password_hash=auth_utils.hash_password(temporary_password),
        role=UserRole.TRAINEE,
    )
    _apply_batch_profile(new_user, batch)

    db.add(new_user)
    db.flush()
    _sync_trainee_to_supabase_or_raise(db, new_user)

    if new_user not in batch.users:
        batch.users.append(new_user)

    return new_user


def _parse_bulk_upload_rows(file_bytes: bytes, filename: str) -> List[Dict[str, str]]:
    normalized_filename = (filename or "").lower()

    header_aliases = {
        "email": "email",
        "email_address": "email",
        "full_name": "full_name",
        "fullname": "full_name",
        "role": "role",
        "password": "password",
        "wave_batch": "batch_lookup",
        "wave_or_batch": "batch_lookup",
        "batch": "batch_lookup",
        "batch_name": "batch_lookup",
        "batch_number": "batch_lookup",
        "wave": "batch_lookup",
        "wave_number": "batch_lookup",
    }

    def map_headers(headers: List[str]) -> List[str]:
        mapped = []
        for header in headers:
            normalized = _normalize_header(header)
            mapped.append(header_aliases.get(normalized, normalized))
        return mapped

    if normalized_filename.endswith(".csv"):
        decoded = file_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(decoded))
        rows = list(reader)
    elif normalized_filename.endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file type. Upload a .xlsx or .csv file.",
        )

    if not rows:
        raise HTTPException(status_code=400, detail="The uploaded file is empty")

    headers = map_headers([_normalize_text_value(cell) for cell in rows[0]])
    parsed_rows: List[Dict[str, str]] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue

        record: Dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = _normalize_text_value(row[index] if index < len(row) else "")
        parsed_rows.append(record)

    return parsed_rows


# ==================== Pydantic Models ====================


class BatchCreate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    wave_number: Optional[int] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    lob: Optional[str] = None
    is_active: bool = True


class BatchUserAssignment(BaseModel):
    user_ids: List[str]  # IDs of trainees to add to batch


class CourseCreate(BaseModel):
    name: str
    description: Optional[str] = None
    duration_minutes: Optional[int] = None
    scenario_ids: Optional[List[str]] = None
    difficulty: Optional[str] = None


class CourseAssignmentCreate(BaseModel):
    course_id: str
    batch_id: Optional[str] = None
    user_id: Optional[str] = None  # For individual assignment
    is_mandatory: bool = True
    due_date: Optional[datetime] = None


class FeedbackCreate(BaseModel):
    practice_session_id: str
    feedback_type: FeedbackType
    content: str
    recommended_module_id: Optional[str] = None
    recommended_exercises: Optional[List[str]] = None


class MicrolearningTopicCategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None


class MicrolearningModuleCreate(BaseModel):
    title: str
    description: Optional[str] = None
    category: FeedbackType
    module_type: str = "quiz"
    duration_minutes: int = 2
    passing_score: int = 75
    skill_focus: Optional[str] = None
    content_url: Optional[str] = None
    content_data: Optional[Dict[str, Any]] = None
    difficulty: ScenarioDifficulty = ScenarioDifficulty.BASIC
    topic_category_id: Optional[str] = None


class MicrolearningAssignmentCreate(BaseModel):
    module_ids: List[str]
    batch_id: Optional[str] = None
    user_id: Optional[str] = None
    due_date: Optional[datetime] = None
    notes: Optional[str] = None
    is_mandatory: bool = True


class SessionVerification(BaseModel):
    is_verified: bool = True
    corrected_transcription: Optional[str] = None


class TraineeCreate(BaseModel):
    email: str
    full_name: str
    role: UserRole = UserRole.TRAINEE
    password: Optional[str] = None
    batch_id: str


class TraineeUpdate(BaseModel):
    email: str
    full_name: str
    role: UserRole = UserRole.TRAINEE
    batch_id: str


class TraineeStatusUpdate(BaseModel):
    is_active: bool


# ==================== Helper Functions ====================


def verify_trainer(
    current_user: User = Depends(auth_utils.get_current_user),
) -> User:
    """Verify that current user is trainer"""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(status_code=403, detail="Trainer access required")
    return current_user


# ==================== Batch Management ====================


@router.post("/batches", response_model=dict)
async def create_batch(
    batch_data: BatchCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Create a new training batch"""
    if batch_data.start_date and batch_data.end_date and batch_data.end_date < batch_data.start_date:
        raise HTTPException(status_code=400, detail="End date cannot be earlier than start date")

    batch_name = _derive_batch_name(batch_data.name, batch_data.wave_number)

    existing_name = (
        db.query(Batch)
        .filter(
            Batch.created_by == current_user.id,
            func.lower(Batch.name) == batch_name.lower(),
        )
        .first()
    )
    if existing_name:
        raise HTTPException(status_code=400, detail="Batch name already exists")

    if batch_data.wave_number is not None:
        existing_wave = (
            db.query(Batch)
            .filter(
                Batch.created_by == current_user.id,
                Batch.wave_number == batch_data.wave_number,
            )
            .first()
        )
        if existing_wave:
            raise HTTPException(status_code=400, detail="Batch number already exists")

    new_batch = Batch(
        name=batch_name,
        description=_normalize_text_value(batch_data.description) or None,
        wave_number=batch_data.wave_number,
        start_date=batch_data.start_date,
        end_date=batch_data.end_date,
        lob=_normalize_text_value(batch_data.lob) or None,
        is_active=batch_data.is_active,
        created_by=current_user.id,
    )

    db.add(new_batch)
    db.commit()
    db.refresh(new_batch)

    return {
        "status": "created",
        **_serialize_trainer_batch_summary(new_batch),
    }


@router.get("/batches")
async def list_batches(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
    skip: int = 0,
    limit: int = 50,
):
    """List all batches created by this trainer"""
    batches = (
        db.query(Batch)
        .filter(Batch.created_by == current_user.id)
        .order_by(Batch.wave_number.is_(None), Batch.wave_number.asc(), Batch.name.asc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    return {
        "count": len(batches),
        "batches": [_serialize_trainer_batch_summary(batch) for batch in batches],
    }


@router.get("/batches/{batch_id}")
async def get_batch(
    batch_id: str, current_user: Any = Depends(verify_trainer), db: Session = Depends()
):
    """Get detailed batch information"""
    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)

    return {
        "id": batch.id,
        "name": batch.name,
        "description": batch.description,
        "wave_number": batch.wave_number,
        "start_date": batch.start_date,
        "end_date": batch.end_date,
        "lob": batch.lob,
        "users": [
            {
                "id": u.id,
                "email": u.email,
                "full_name": u.full_name,
                "language_dialect": u.language_dialect,
            }
            for u in batch.users
        ],
        "course_assignments": [
            {
                "id": ca.id,
                "course_id": ca.course_id,
                "course_name": ca.course.name if ca.course else None,
                "assigned_at": ca.assigned_at,
                "due_date": ca.due_date,
            }
            for ca in batch.course_assignments
        ],
        "created_at": batch.created_at,
    }


@router.put("/batches/{batch_id}")
async def update_batch(
    batch_id: str,
    batch_data: BatchCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Update an existing trainer-owned batch."""
    if batch_data.start_date and batch_data.end_date and batch_data.end_date < batch_data.start_date:
        raise HTTPException(status_code=400, detail="End date cannot be earlier than start date")

    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)
    batch_name = _derive_batch_name(batch_data.name, batch_data.wave_number)

    existing_name = (
        db.query(Batch)
        .filter(
            Batch.created_by == current_user.id,
            func.lower(Batch.name) == batch_name.lower(),
            Batch.id != batch.id,
        )
        .first()
    )
    if existing_name:
        raise HTTPException(status_code=400, detail="Batch name already exists")

    if batch_data.wave_number is not None:
        existing_wave = (
            db.query(Batch)
            .filter(
                Batch.created_by == current_user.id,
                Batch.wave_number == batch_data.wave_number,
                Batch.id != batch.id,
            )
            .first()
        )
        if existing_wave:
            raise HTTPException(status_code=400, detail="Batch number already exists")

    batch.name = batch_name
    batch.description = _normalize_text_value(batch_data.description) or None
    batch.wave_number = batch_data.wave_number
    batch.start_date = batch_data.start_date
    batch.end_date = batch_data.end_date
    batch.lob = _normalize_text_value(batch_data.lob) or None
    batch.is_active = batch_data.is_active

    for trainee in [user for user in batch.users if user.role == UserRole.TRAINEE]:
        _apply_batch_profile(trainee, batch)

    db.commit()
    db.refresh(batch)

    return {
        "status": "updated",
        "batch": _serialize_trainer_batch_summary(batch),
    }


@router.delete("/batches/{batch_id}")
async def delete_batch(
    batch_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Delete a trainer-owned batch and safely clean up linked trainer data."""
    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)

    batch_label = batch.name
    linked_users = list(batch.users or [])
    trainee_count = len([user for user in linked_users if user.role == UserRole.TRAINEE])

    batch_course_assignments = (
        db.query(CourseAssignment)
        .filter(
            CourseAssignment.batch_id == batch.id,
            CourseAssignment.assigned_by == current_user.id,
        )
        .all()
    )
    batch_microlearning_assignments = (
        db.query(MicrolearningAssignment)
        .filter(
            MicrolearningAssignment.batch_id == batch.id,
            MicrolearningAssignment.assigned_by == current_user.id,
        )
        .all()
    )

    for user in linked_users:
        if user in batch.users:
            batch.users.remove(user)
        if user.role == UserRole.TRAINEE:
            _refresh_trainee_profile_from_batches(user)

    for assignment in batch_microlearning_assignments:
        assignment.batch_id = None

    deleted_course_assignments = len(batch_course_assignments)
    for assignment in batch_course_assignments:
        db.delete(assignment)

    preserved_microlearning_assignments = len(batch_microlearning_assignments)
    db.delete(batch)
    db.commit()

    return {
        "status": "deleted",
        "batch_id": batch_id,
        "name": batch_label,
        "removed_trainees": trainee_count,
        "deleted_course_assignments": deleted_course_assignments,
        "preserved_microlearning_assignments": preserved_microlearning_assignments,
    }


@router.post("/batches/{batch_id}/users")
async def assign_users_to_batch(
    batch_id: str,
    assignment: BatchUserAssignment,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Add trainees to a batch"""
    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)
    requested_ids = [user_id for user_id in dict.fromkeys(assignment.user_ids) if user_id]
    if not requested_ids:
        raise HTTPException(status_code=400, detail="Select at least one trainee to add")

    users = (
        db.query(User)
        .filter(
            User.id.in_(requested_ids),
            User.role == UserRole.TRAINEE,
            User.is_active.is_(True),
        )
        .all()
    )

    if not users:
        raise HTTPException(status_code=404, detail="No users found")

    added_count = 0
    moved_count = 0
    for user in users:
        trainer_batches = [
            existing_batch for existing_batch in user.batches if existing_batch.created_by == current_user.id
        ]
        moved_from_other_batch = any(existing_batch.id != batch.id for existing_batch in trainer_batches)
        for existing_batch in trainer_batches:
            if existing_batch.id != batch.id and user in existing_batch.users:
                existing_batch.users.remove(user)
        if user not in batch.users:
            batch.users.append(user)
            added_count += 1
        if moved_from_other_batch:
            moved_count += 1
        _apply_batch_profile(user, batch)

    db.commit()

    return {
        "status": "updated",
        "batch_id": batch_id,
        "added_users": added_count,
        "moved_users": moved_count,
        "total_users": len([user for user in batch.users if user.role == UserRole.TRAINEE]),
        "trainees": [_serialize_trainee(user, batch=batch) for user in users],
    }


@router.delete("/batches/{batch_id}/users/{user_id}")
async def remove_user_from_batch(
    batch_id: str,
    user_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Remove a trainee from a batch"""
    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)

    user = (
        db.query(User)
        .filter(
            User.id == user_id,
            User.role == UserRole.TRAINEE,
            User.is_active.is_(True),
        )
        .first()
    )
    if not user:
        raise HTTPException(status_code=404, detail="Trainee not found")

    if user in batch.users:
        batch.users.remove(user)
        _refresh_trainee_profile_from_batches(user)
        db.commit()

    return {"status": "removed", "batch_id": batch_id, "user_id": user_id}


@router.post("/trainees", response_model=dict)
async def create_trainee(
    trainee_data: TraineeCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Create a trainee account and assign it to one of the trainer's batches."""
    if trainee_data.role != UserRole.TRAINEE:
        raise HTTPException(
            status_code=400,
            detail="Trainer user management only creates Trainee accounts.",
        )
    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=trainee_data.batch_id)

    temporary_password = trainee_data.password or DEFAULT_TRAINEE_PASSWORD
    try:
        new_user = _create_trainee_account(
            db,
            email=trainee_data.email,
            full_name=trainee_data.full_name,
            batch=batch,
            password=temporary_password,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except HTTPException:
        db.rollback()
        raise

    db.commit()
    db.refresh(new_user)

    return {
        "status": "created",
        "id": new_user.id,
        "email": new_user.email,
        "full_name": new_user.full_name,
        "role": new_user.role,
        "temporary_password": temporary_password,
        "batch": _serialize_batch(batch),
    }


@router.get("/trainees")
async def list_trainees(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """List active trainees assigned to batches created by the current trainer."""
    trainees = (
        db.query(User, Batch)
        .join(batch_user_association, batch_user_association.c.user_id == User.id)
        .join(Batch, Batch.id == batch_user_association.c.batch_id)
        .filter(
            User.role == UserRole.TRAINEE,
            User.is_active.is_(True),
            Batch.created_by == current_user.id,
        )
        .order_by(Batch.wave_number.asc(), User.full_name.asc())
        .all()
    )

    trainee_lookup: Dict[str, Dict[str, Any]] = {}
    for user, batch in trainees:
        entry = trainee_lookup.setdefault(user.id, {"user": user, "batches": []})
        if all(existing_batch.id != batch.id for existing_batch in entry["batches"]):
            entry["batches"].append(batch)

    trainee_rows = [
        _serialize_trainee(
            entry["user"],
            batches=_sort_batches_for_display(entry["batches"]),
        )
        for entry in trainee_lookup.values()
    ]
    trainee_rows.sort(
        key=lambda trainee: (
            trainee["batch"]["wave_number"] if trainee.get("batch") and trainee["batch"].get("wave_number") is not None else 10**6,
            trainee["full_name"].lower(),
            trainee["email"].lower(),
        )
    )

    return {"count": len(trainee_rows), "trainees": trainee_rows}


@router.get("/trainees/registry")
async def list_registered_trainees(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """List every active trainee account so trainers can add existing users into their class batches."""
    trainer_batches = (
        db.query(Batch)
        .filter(Batch.created_by == current_user.id)
        .order_by(Batch.wave_number.is_(None), Batch.wave_number.asc(), Batch.name.asc())
        .all()
    )
    trainer_batch_ids = {batch.id for batch in trainer_batches}

    trainees = (
        db.query(User)
        .filter(User.role == UserRole.TRAINEE)
        .order_by(User.is_active.desc(), User.full_name.asc(), User.email.asc())
        .all()
    )

    trainee_rows = []
    for trainee in trainees:
        all_batches = _sort_batches_for_display(list(trainee.batches or []))
        trainer_owned_batches = [
            existing_batch for existing_batch in all_batches if existing_batch.id in trainer_batch_ids
        ]
        current_trainer_batch = trainer_owned_batches[0] if trainer_owned_batches else None
        serialized = _serialize_trainee(trainee, batches=all_batches)
        serialized.update(
            {
                "is_in_my_class": bool(trainer_owned_batches),
                "current_trainer_batch_id": current_trainer_batch.id if current_trainer_batch else None,
                "current_trainer_batch_name": (
                    _build_batch_department_label(current_trainer_batch)
                    if current_trainer_batch
                    else None
                ),
            }
        )
        trainee_rows.append(serialized)

    return {"count": len(trainee_rows), "trainees": trainee_rows}


@router.get("/all-trainees")
async def list_all_trainees_in_system(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """
    List ALL trainee accounts in the system (both active and inactive).
    Used for trainee status management. Trainers can activate/deactivate any trainee.
    """
    trainees = (
        db.query(User)
        .filter(User.role == UserRole.TRAINEE)
        .order_by(User.is_active.desc(), User.full_name.asc(), User.email.asc())
        .all()
    )

    trainee_rows = []
    for trainee in trainees:
        batches = _sort_batches_for_display(list(trainee.batches or []))
        row = {
            "id": trainee.id,
            "full_name": trainee.full_name,
            "email": trainee.email,
            "is_active": trainee.is_active,
            "department": trainee.department,
            "batch": _build_batch_row(batches[0]) if batches else None,
            "batches": [_build_batch_row(b) for b in batches],
            "created_at": trainee.created_at,
        }
        trainee_rows.append(row)

    return {"count": len(trainee_rows), "trainees": trainee_rows}


@router.put("/trainees/{trainee_id}", response_model=dict)
async def update_trainee(
    trainee_id: str,
    trainee_data: TraineeUpdate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Update a trainee profile and reassign them to one of the trainer's batches."""
    if trainee_data.role != UserRole.TRAINEE:
        raise HTTPException(
            status_code=400,
            detail="Trainer user management only updates Trainee accounts.",
        )

    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=trainee_data.batch_id)
    trainee = (
        db.query(User)
        .filter(
            User.id == trainee_id,
            User.role == UserRole.TRAINEE,
            User.is_active.is_(True),
        )
        .first()
    )

    if not trainee:
        raise HTTPException(status_code=404, detail="Trainee not found")

    trainer_batches = [existing_batch for existing_batch in trainee.batches if existing_batch.created_by == current_user.id]
    if not trainer_batches:
        raise HTTPException(status_code=404, detail="Trainee not found in your batch list")

    normalized_email = _normalize_email(trainee_data.email)
    normalized_name = _normalize_text_value(trainee_data.full_name)

    if not normalized_email:
        raise HTTPException(status_code=400, detail="Email address is required")
    if not normalized_name:
        raise HTTPException(status_code=400, detail="Full name is required")

    existing_user = (
        db.query(User)
        .filter(
            func.lower(User.email) == normalized_email,
            User.id != trainee.id,
        )
        .first()
    )
    if existing_user:
        raise HTTPException(status_code=400, detail=f"Email {normalized_email} already exists")

    trainee.email = normalized_email
    trainee.full_name = normalized_name
    _apply_batch_profile(trainee, batch)

    for existing_batch in trainer_batches:
        if existing_batch.id != batch.id and trainee in existing_batch.users:
            existing_batch.users.remove(trainee)

    if trainee not in batch.users:
        batch.users.append(trainee)

    try:
        _sync_trainee_to_supabase_or_raise(db, trainee)
    except HTTPException:
        db.rollback()
        raise
    db.commit()
    db.refresh(trainee)

    return {
        "status": "updated",
        "trainee": _serialize_trainee(trainee, batch),
    }


@router.put("/trainees/{trainee_id}/status", response_model=dict)
async def update_trainee_status(
    trainee_id: str,
    status_data: TraineeStatusUpdate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """
    Update a trainee's active/inactive status.
    Trainers can activate/deactivate any trainee in the system.
    If deactivating, removes the trainee from all batches.
    """
    trainee = (
        db.query(User)
        .filter(
            User.id == trainee_id,
            User.role == UserRole.TRAINEE,
        )
        .first()
    )

    if not trainee:
        raise HTTPException(status_code=404, detail="Trainee not found")

    # If deactivating a trainee, remove them from all batches
    if not status_data.is_active:
        # Remove from all batches, regardless of ownership
        for batch in list(trainee.batches):
            if trainee in batch.users:
                batch.users.remove(trainee)

    trainee.is_active = status_data.is_active
    db.commit()
    db.refresh(trainee)

    return {
        "status": "updated",
        "trainee": {
            "id": trainee.id,
            "email": trainee.email,
            "full_name": trainee.full_name,
            "is_active": trainee.is_active,
        },
    }


@router.get("/trainees/bulk-upload-template")
async def download_trainee_bulk_upload_template(
    current_user: Any = Depends(verify_trainer),
    format: str = Query("csv"),
):
    """Download the trainee bulk upload template as CSV or Excel."""
    headers = ["Email Address", "Full Name", "Role", "Password", "Wave/Batch"]

    template_format = (format or "csv").strip().lower()
    if template_format == "xlsx":
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Trainees"
        worksheet.append(headers)

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Notes"])
        instructions.append(["Email Address", "Required and must be unique"])
        instructions.append(["Full Name", "Required"])
        instructions.append(["Role", "Use trainee"])
        instructions.append(["Password", f"Default password is always {DEFAULT_TRAINEE_PASSWORD}"])
        instructions.append(["Wave/Batch", "Use an existing active trainer batch name or batch number"])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{TRAINER_BULK_UPLOAD_TEMPLATE}"'
            },
        )

    if template_format != "csv":
        raise HTTPException(status_code=400, detail="Template format must be csv or xlsx")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="trainer-trainee-bulk-upload-template.csv"'
        },
    )


@router.post("/trainees/bulk-upload")
async def bulk_upload_trainees(
    file: UploadFile = File(...),
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Bulk create trainee accounts from an Excel or CSV file."""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    rows = _parse_bulk_upload_rows(content, file.filename or "")
    if not rows:
        raise HTTPException(status_code=400, detail="No trainee rows found in the uploaded file")

    created_rows = []
    errors = []

    for index, row in enumerate(rows, start=2):
        email = _normalize_text_value(row.get("email"))
        full_name = _normalize_text_value(row.get("full_name"))
        role = _normalize_text_value(row.get("role")).lower() or UserRole.TRAINEE.value
        batch_lookup = _normalize_text_value(row.get("batch_lookup"))

        if role != UserRole.TRAINEE.value:
            errors.append(f"Row {index}: Role must be trainee")
            continue
        if not email or not full_name:
            errors.append(f"Row {index}: Email address and full name are required")
            continue
        if not batch_lookup:
            errors.append(f"Row {index}: Wave/Batch is required")
            continue

        batch = _resolve_batch_lookup(
            db,
            trainer_id=current_user.id,
            batch_lookup=batch_lookup,
        )
        if not batch:
            errors.append(
                f"Row {index}: No trainer batch matches '{batch_lookup}'. Create the batch first."
            )
            continue

        try:
            with db.begin_nested():
                new_user = _create_trainee_account(
                    db,
                    email=email,
                    full_name=full_name,
                    batch=batch,
                    password=DEFAULT_TRAINEE_PASSWORD,
                )
                created_rows.append(
                    _serialize_trainee(new_user, batch)
                )
        except ValueError as exc:
            errors.append(f"Row {index}: {exc}")
        except HTTPException as exc:
            errors.append(f"Row {index}: {exc.detail}")

    db.commit()

    return {
        "status": "completed",
        "created": len(created_rows),
        "temporary_password": DEFAULT_TRAINEE_PASSWORD,
        "trainees": created_rows,
        "errors": errors,
    }


# ==================== Microlearning Management ====================


@router.get("/microlearning-topic-categories")
async def list_microlearning_topic_categories(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """List trainer-managed microlearning topic categories."""
    categories = (
        db.query(MicrolearningTopicCategory)
        .filter(
            MicrolearningTopicCategory.created_by == current_user.id,
            MicrolearningTopicCategory.is_active == True,
        )
        .order_by(MicrolearningTopicCategory.name.asc())
        .all()
    )
    return {
        "count": len(categories),
        "categories": [serialize_topic_category(category) for category in categories],
    }


@router.post("/microlearning-topic-categories")
async def create_microlearning_topic_category(
    payload: MicrolearningTopicCategoryCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Create a trainer-owned microlearning topic category."""
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Category name is required")

    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    existing = (
        db.query(MicrolearningTopicCategory)
        .filter(
            MicrolearningTopicCategory.created_by == current_user.id,
            MicrolearningTopicCategory.slug == slug,
        )
        .first()
    )
    if existing and existing.is_active:
        raise HTTPException(status_code=400, detail="Category name already exists")

    if existing:
        existing.name = name
        existing.description = (payload.description or "").strip() or None
        existing.is_active = True
        db.commit()
        db.refresh(existing)
        return {"status": "restored", "category": serialize_topic_category(existing)}

    category = MicrolearningTopicCategory(
        name=name,
        slug=slug,
        description=(payload.description or "").strip() or None,
        created_by=current_user.id,
        is_active=True,
    )
    db.add(category)
    db.commit()
    db.refresh(category)
    return {"status": "created", "category": serialize_topic_category(category)}


@router.put("/microlearning-topic-categories/{category_id}")
async def update_microlearning_topic_category(
    category_id: str,
    payload: MicrolearningTopicCategoryCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Update a trainer-owned microlearning topic category."""
    category = _get_trainer_microlearning_topic_category(
        db,
        trainer_id=current_user.id,
        category_id=category_id,
    )
    name = (payload.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Category name is required")

    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    duplicate = (
        db.query(MicrolearningTopicCategory)
        .filter(
            MicrolearningTopicCategory.created_by == current_user.id,
            MicrolearningTopicCategory.slug == slug,
            MicrolearningTopicCategory.id != category.id,
            MicrolearningTopicCategory.is_active == True,
        )
        .first()
    )
    if duplicate:
        raise HTTPException(status_code=400, detail="Category name already exists")

    category.name = name
    category.slug = slug
    category.description = (payload.description or "").strip() or None
    db.commit()
    db.refresh(category)
    return {"status": "updated", "category": serialize_topic_category(category)}


@router.delete("/microlearning-topic-categories/{category_id}")
async def delete_microlearning_topic_category(
    category_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Soft-delete a trainer-owned microlearning topic category."""
    category = _get_trainer_microlearning_topic_category(
        db,
        trainer_id=current_user.id,
        category_id=category_id,
    )
    linked_modules = (
        db.query(MicrolearningModule)
        .filter(
            MicrolearningModule.created_by == current_user.id,
            MicrolearningModule.topic_category_id == category.id,
        )
        .all()
    )

    reassigned_module_count = 0
    for module in linked_modules:
        if module.is_active:
            reassigned_module_count += 1
        module.topic_category_id = None
        module.topic_category = None

    category.is_active = False
    db.commit()
    return {
        "status": "deleted",
        "category_id": category.id,
        "reassigned_module_count": reassigned_module_count,
    }


@router.post("/microlearning-assets/upload")
async def upload_microlearning_asset(
    file: UploadFile = File(...),
    module_id: Optional[str] = Form(None),
    module_type: Optional[str] = Form(None),
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Upload a trainer asset for video, image, audio, or infographic content."""
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    upload_result = _upload_microlearning_asset(
        db=db,
        trainer_id=current_user.id,
        file_bytes=file_bytes,
        filename=file.filename or "microlearning-asset.bin",
        content_type=file.content_type,
        module_id=module_id,
        module_type=module_type,
    )
    return {
        "status": "uploaded",
        **upload_result,
        "filename": file.filename,
    }


@router.post("/microlearning-library/seed-bpo-pack")
async def seed_bpo_microlearning_pack(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Legacy seed endpoint retained only to block reintroducing sample content."""
    raise HTTPException(
        status_code=410,
        detail="Default microlearning seed packs are disabled. Categories and modules must be created by the trainer.",
    )


@router.get("/microlearning-reports/overview")
async def get_microlearning_report_overview(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Return trainer-facing microlearning analytics grouped by batch and trainee."""
    return _build_microlearning_report_overview(db, trainer_id=current_user.id)


@router.get("/microlearning-assessment-methods")
async def list_microlearning_assessment_methods(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Trainer microlearning methods have been removed from the authoring workflow."""
    return {"count": 0, "methods": []}


@router.post("/microlearning-assessment-methods/seed-examples")
async def seed_microlearning_assessment_examples(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Legacy seed endpoint retained only to block reintroducing sample lessons."""
    raise HTTPException(
        status_code=410,
        detail="Example microlearning lessons are disabled. Modules must be created by the trainer.",
    )


@router.get("/microlearning-modules")
async def list_microlearning_modules(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """List active microlearning modules created by the current trainer."""
    modules = (
        db.query(MicrolearningModule)
        .options(
            selectinload(MicrolearningModule.assessment_method),
            selectinload(MicrolearningModule.topic_category),
        )
        .filter(
            MicrolearningModule.created_by == current_user.id,
            MicrolearningModule.is_active == True,
        )
        .order_by(MicrolearningModule.created_at.desc())
        .all()
    )

    did_backfill = False
    for module in modules:
        did_backfill = ensure_module_exercises(module) or did_backfill

    if did_backfill:
        db.commit()

    assignment_counts = _get_microlearning_assignment_counts(
        db,
        module_ids=[module.id for module in modules],
    )

    return {
        "count": len(modules),
        "modules": [
            serialize_microlearning_module(
                module,
                assignment_count=assignment_counts.get(module.id, 0),
            )
            for module in modules
        ],
    }


@router.post("/microlearning-modules")
async def create_microlearning_module(
    payload: MicrolearningModuleCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Create a trainer-owned microlearning module using trainer-authored content only."""
    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="Module title is required")
    if payload.duration_minutes <= 0:
        raise HTTPException(status_code=400, detail="Duration must be greater than zero")
    if payload.passing_score < 1 or payload.passing_score > 100:
        raise HTTPException(status_code=400, detail="Passing score must be between 1 and 100")

    module_type = normalize_module_type(payload.module_type)
    if module_type == "case_study":
        raise HTTPException(
            status_code=400,
            detail="Legacy case study modules can no longer be created. Use the Audio Lesson format instead.",
        )
    feedback_category_value = _feedback_category_api_value(payload.category)
    feedback_category_db_label = _feedback_category_db_label(payload.category)

    topic_category: Optional[MicrolearningTopicCategory] = None
    if payload.topic_category_id:
        topic_category = _get_trainer_microlearning_topic_category(
            db,
            trainer_id=current_user.id,
            category_id=payload.topic_category_id,
        )

    content_data = dict(payload.content_data or {})
    if payload.content_url and not content_data.get("asset_url"):
        content_data["asset_url"] = payload.content_url
    if content_data.get("asset_record_id"):
        content_data["storage_backend"] = content_data.get("storage_backend") or "supabase_postgres"
    normalized_asset_bucket, normalized_asset_path = _normalize_supabase_asset_reference(content_data)
    if normalized_asset_path:
        content_data["asset_storage_path"] = normalized_asset_path
        content_data["asset_bucket"] = normalized_asset_bucket or get_supabase_client().microlearning_bucket_name
    if content_data.get("asset_storage_path"):
        content_data["signed_url_required"] = bool(content_data.get("signed_url_required", True))

    exercises = build_type_specific_exercises(
        module_type,
        content_data,
        title=title,
        skill_focus=payload.skill_focus,
    )

    module = MicrolearningModule(
        title=title,
        description=(payload.description or "").strip() or None,
        category=feedback_category_db_label,
        type=module_type,
        duration_minutes=payload.duration_minutes,
        content_data=content_data,
        passing_score=payload.passing_score,
        skill_focus=(payload.skill_focus or "").strip() or None,
        content_url=(payload.content_url or "").strip() or None,
        difficulty=payload.difficulty,
        exercises=exercises,
        assessment_method_id=None,
        topic_category_id=topic_category.id if topic_category else None,
        created_by=current_user.id,
    )
    if topic_category:
        module.topic_category = topic_category
    db.add(module)
    db.commit()
    db.refresh(module)
    if topic_category:
        module.topic_category = topic_category

    return {
        "status": "created",
        "module": serialize_microlearning_module(module, assignment_count=0),
    }


@router.put("/microlearning-modules/{module_id}")
async def update_microlearning_module(
    module_id: str,
    payload: MicrolearningModuleCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Update a trainer-owned microlearning module stored in the database."""
    module = _get_trainer_microlearning_module(
        db,
        trainer_id=current_user.id,
        module_id=module_id,
    )

    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="Module title is required")
    if payload.duration_minutes <= 0:
        raise HTTPException(status_code=400, detail="Duration must be greater than zero")
    if payload.passing_score < 1 or payload.passing_score > 100:
        raise HTTPException(status_code=400, detail="Passing score must be between 1 and 100")

    next_skill_focus = (payload.skill_focus or "").strip() or None
    next_description = (payload.description or "").strip() or None
    next_content_url = (payload.content_url or "").strip() or None
    previous_content_data = dict(module.content_data or {})
    module_type = normalize_module_type(payload.module_type)
    if module_type == "case_study" and normalize_module_type(module.type) != "case_study":
        raise HTTPException(
            status_code=400,
            detail="Legacy case study modules can no longer be created. Use the Audio Lesson format instead.",
        )
    feedback_category_value = _feedback_category_api_value(payload.category)
    feedback_category_db_label = _feedback_category_db_label(payload.category)

    topic_category: Optional[MicrolearningTopicCategory] = None
    if payload.topic_category_id:
        topic_category = _get_trainer_microlearning_topic_category(
            db,
            trainer_id=current_user.id,
            category_id=payload.topic_category_id,
        )

    next_content_data = dict(payload.content_data or {})
    if next_content_url and not next_content_data.get("asset_url"):
        next_content_data["asset_url"] = next_content_url
    if next_content_data.get("asset_record_id"):
        next_content_data["storage_backend"] = next_content_data.get("storage_backend") or "supabase_postgres"
    normalized_asset_bucket, normalized_asset_path = _normalize_supabase_asset_reference(next_content_data)
    if normalized_asset_path:
        next_content_data["asset_storage_path"] = normalized_asset_path
        next_content_data["asset_bucket"] = normalized_asset_bucket or get_supabase_client().microlearning_bucket_name
    if next_content_data.get("asset_storage_path"):
        next_content_data["signed_url_required"] = bool(next_content_data.get("signed_url_required", True))
    cleanup_storage_target, cleanup_asset_record_id = _prepare_replaced_module_asset_cleanup(
        previous_content_data=previous_content_data,
        next_content_data=next_content_data,
    )

    should_regenerate_exercises = any(
        [
            module.title != title,
            str(getattr(module, "category", "") or "").strip().lower() != feedback_category_value,
            normalize_module_type(module.type) != module_type,
            module.skill_focus != next_skill_focus,
            module.topic_category_id != (topic_category.id if topic_category else None),
            (module.content_data or {}) != next_content_data,
            not module.exercises,
        ]
    )
    has_existing_assignments = (
        db.query(MicrolearningAssignment.id)
        .filter(MicrolearningAssignment.module_id == module.id)
        .first()
        is not None
    )

    module.title = title
    module.description = next_description
    module.category = feedback_category_db_label
    module.type = module_type
    module.duration_minutes = payload.duration_minutes
    module.content_data = next_content_data
    module.passing_score = payload.passing_score
    module.skill_focus = next_skill_focus
    module.content_url = next_content_url
    module.difficulty = payload.difficulty
    module.assessment_method_id = None
    module.assessment_method = None
    module.topic_category_id = topic_category.id if topic_category else None
    module.topic_category = topic_category

    exercises_regenerated = False
    exercises_locked = False
    if should_regenerate_exercises:
        if has_existing_assignments:
            exercises_locked = True
        else:
            module.exercises = build_type_specific_exercises(
                module_type,
                next_content_data,
                title=title,
                skill_focus=next_skill_focus,
            )
            exercises_regenerated = True

    if cleanup_asset_record_id:
        (
            db.query(MicrolearningUploadedAsset)
            .filter(MicrolearningUploadedAsset.id == cleanup_asset_record_id)
            .delete(synchronize_session=False)
        )

    db.commit()
    db.refresh(module)
    module.topic_category = topic_category
    assignment_count = _get_microlearning_assignment_counts(db, module_ids=[module.id]).get(module.id, 0)
    if cleanup_storage_target:
        bucket_name, storage_path = cleanup_storage_target
        get_supabase_client().delete_storage_object(
            bucket_name=bucket_name,
            path=storage_path,
        )

    return {
        "status": "updated",
        "module": serialize_microlearning_module(module, assignment_count=assignment_count),
        "exercises_regenerated": exercises_regenerated,
        "exercises_locked": exercises_locked,
    }


@router.delete("/microlearning-modules/{module_id}")
async def delete_microlearning_module(
    module_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Permanently delete a trainer-owned microlearning module and its dependent records."""
    module = _get_trainer_microlearning_module(
        db,
        trainer_id=current_user.id,
        module_id=module_id,
    )
    supabase_client = get_supabase_client()

    try:
        delete_summary = delete_microlearning_module_and_dependencies(
            db,
            module=module,
            supabase_client=supabase_client,
        )
    except MicrolearningDeleteError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc

    for trainee_id in delete_summary.impacted_trainee_ids:
        await live_update_manager.broadcast(
            f"trainee:{trainee_id}",
            {
                "type": "microlearning_module_deleted",
                "module_id": delete_summary.module_id,
                "title": delete_summary.module_title,
                "deleted_assignments": delete_summary.deleted_assignment_count,
            },
        )

    return {
        "status": "deleted",
        "module_id": delete_summary.module_id,
        "title": delete_summary.module_title,
        "deleted_assignments": delete_summary.deleted_assignment_count,
        "deleted_certificates": delete_summary.deleted_certificate_count,
        "deleted_storage_count": delete_summary.deleted_storage_count,
        "impacted_trainee_ids": delete_summary.impacted_trainee_ids,
        "impacted_batch_ids": delete_summary.impacted_batch_ids,
    }


@router.get("/microlearning-assignments")
async def list_microlearning_assignments(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """List microlearning assignments created by the current trainer."""
    assignments = (
        db.query(MicrolearningAssignment)
        .options(
            selectinload(MicrolearningAssignment.module).selectinload(MicrolearningModule.assessment_method),
            selectinload(MicrolearningAssignment.trainee).selectinload(User.batches),
            selectinload(MicrolearningAssignment.trainer),
            selectinload(MicrolearningAssignment.batch),
            selectinload(MicrolearningAssignment.certificate),
        )
        .filter(MicrolearningAssignment.assigned_by == current_user.id)
        .order_by(MicrolearningAssignment.assigned_at.desc())
        .all()
    )
    assignments = filter_current_assignments(assignments)

    did_update = False
    serialized = []
    for assignment in assignments:
        did_update = ensure_module_exercises(assignment.module) or did_update
        before = (
            assignment.status,
            assignment.completion_percentage,
            assignment.completed_exercises,
            assignment.completed_at,
        )
        refresh_assignment_progress(assignment)
        after = (
            assignment.status,
            assignment.completion_percentage,
            assignment.completed_exercises,
            assignment.completed_at,
        )
        did_update = before != after or did_update
        serialized.append(serialize_assignment_summary(assignment))

    if did_update:
        db.commit()

    return {
        "count": len(serialized),
        "assignments": serialized,
    }


@router.post("/microlearning-assignments")
async def assign_microlearning_modules(
    payload: MicrolearningAssignmentCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Assign selected modules to a batch or a specific trainee."""
    module_ids = [module_id for module_id in dict.fromkeys(payload.module_ids) if module_id]
    if not module_ids:
        raise HTTPException(status_code=400, detail="Select at least one module to assign")
    if bool(payload.batch_id) == bool(payload.user_id):
        raise HTTPException(
            status_code=400,
            detail="Specify either a batch_id or a user_id for the assignment.",
        )

    modules = (
        db.query(MicrolearningModule)
        .filter(
            MicrolearningModule.id.in_(module_ids),
            MicrolearningModule.created_by == current_user.id,
            MicrolearningModule.is_active == True,
        )
        .all()
    )
    module_lookup = {module.id: module for module in modules}
    missing_modules = [module_id for module_id in module_ids if module_id not in module_lookup]
    if missing_modules:
        raise HTTPException(status_code=404, detail="One or more modules were not found")

    modules_missing_media = [
        f"{module.title}: {get_module_media_state(module).get('media_status')}"
        for module in modules
        if not get_module_media_state(module).get("media_ready")
    ]
    if modules_missing_media:
        raise HTTPException(
            status_code=400,
            detail="Complete the required media before assigning these modules: "
            + "; ".join(modules_missing_media),
        )

    if payload.batch_id:
        batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=payload.batch_id)
        trainees = [
            user
            for user in batch.users
            if user.role == UserRole.TRAINEE and bool(getattr(user, "is_active", True))
        ]
        if not trainees:
            raise HTTPException(status_code=400, detail="Selected batch has no trainees")
        batch_id = batch.id
    else:
        trainee = _get_trainer_trainee(db, trainer_id=current_user.id, trainee_id=payload.user_id or "")
        trainees = [trainee]
        batch_id = None

    target_trainee_ids = [trainee.id for trainee in trainees]
    assigned_count = 0
    skipped_count = 0
    did_backfill = False

    for trainee in trainees:
        for module_id in module_ids:
            module = module_lookup[module_id]
            did_backfill = ensure_module_exercises(module) or did_backfill

            existing_assignments = (
                db.query(MicrolearningAssignment)
                .options(
                    selectinload(MicrolearningAssignment.module),
                    selectinload(MicrolearningAssignment.batch),
                    selectinload(MicrolearningAssignment.trainee).selectinload(User.batches),
                )
                .filter(
                    MicrolearningAssignment.trainee_id == trainee.id,
                    MicrolearningAssignment.module_id == module.id,
                    MicrolearningAssignment.status.in_(["assigned", "in_progress"]),
                )
                .all()
            )
            if any(assignment_is_current(existing) for existing in existing_assignments):
                skipped_count += 1
                continue

            assignment = MicrolearningAssignment(
                module_id=module.id,
                trainee_id=trainee.id,
                assigned_by=current_user.id,
                batch_id=batch_id,
                due_date=payload.due_date,
                notes=(payload.notes or "").strip() or None,
                is_mandatory=payload.is_mandatory,
                responses={},
            )
            assignment.module = module
            refresh_assignment_progress(assignment)
            db.add(assignment)
            assigned_count += 1

    db.commit()

    for trainee_id in target_trainee_ids:
        await live_update_manager.broadcast(
            f"trainee:{trainee_id}",
            {
                "type": "microlearning_assignments_changed",
                "batch_id": batch_id,
                "module_ids": module_ids,
                "assigned_count": assigned_count,
                "skipped_count": skipped_count,
            },
        )

    return {
        "status": "assigned",
        "assigned_count": assigned_count,
        "skipped_count": skipped_count,
        "backfilled_modules": did_backfill,
    }


@router.delete("/microlearning-assignments/{assignment_id}")
async def delete_microlearning_assignment(
    assignment_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Remove a trainer-owned microlearning assignment and its trainee-facing certificate records."""
    assignment = (
        db.query(MicrolearningAssignment)
        .options(
            selectinload(MicrolearningAssignment.module),
            selectinload(MicrolearningAssignment.trainee),
        )
        .filter(
            MicrolearningAssignment.id == assignment_id,
            MicrolearningAssignment.assigned_by == current_user.id,
        )
        .first()
    )
    if not assignment:
        raise HTTPException(status_code=404, detail="Microlearning assignment not found")

    module_title = assignment.module.title if assignment.module else None
    trainee_name = assignment.trainee.full_name if assignment.trainee else None
    trainee_id = assignment.trainee_id
    linked_certificates: dict[str, CertificateRecord] = {
        certificate.id: certificate
        for certificate in (
            db.query(CertificateRecord)
            .filter(
                CertificateRecord.source_type == "microlearning_assignment",
                CertificateRecord.source_id == assignment.id,
            )
            .all()
        )
    }
    if assignment.certificate_id:
        certificate = (
            db.query(CertificateRecord)
            .filter(CertificateRecord.id == assignment.certificate_id)
            .first()
        )
        if certificate:
            linked_certificates[certificate.id] = certificate

    for certificate in linked_certificates.values():
        db.delete(certificate)

    assignment.certificate_id = None
    db.delete(assignment)
    db.commit()
    sync_trainee_completion_certificates(db, trainee_id)
    return {
        "status": "deleted",
        "assignment_id": assignment_id,
        "module_title": module_title,
        "trainee_name": trainee_name,
        "deleted_certificates": len(linked_certificates),
    }


# ==================== Course Management ====================


@router.post("/courses", response_model=dict)
async def create_course(
    course_data: CourseCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Create a new training course"""
    new_course = Course(
        name=course_data.name,
        description=course_data.description,
        duration_minutes=course_data.duration_minutes,
        scenario_ids=course_data.scenario_ids or [],
        difficulty=course_data.difficulty,
        created_by=current_user.id,
    )

    db.add(new_course)
    db.commit()

    return {"id": new_course.id, "name": new_course.name, "status": "created"}


@router.get("/courses")
async def list_courses(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
    skip: int = 0,
    limit: int = 50,
):
    """List all courses created by this trainer"""
    courses = (
        db.query(Course)
        .filter(Course.created_by == current_user.id)
        .offset(skip)
        .limit(limit)
        .all()
    )

    return {
        "count": len(courses),
        "courses": [
            {
                "id": c.id,
                "name": c.name,
                "duration_minutes": c.duration_minutes,
                "difficulty": c.difficulty,
                "is_published": c.is_published,
                "scenario_count": len(c.scenario_ids),
                "created_at": c.created_at,
            }
            for c in courses
        ],
    }


@router.get("/courses/{course_id}")
async def get_course(
    course_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Get course details"""
    course = _get_trainer_course(db, trainer_id=current_user.id, course_id=course_id)
    return {
        "id": course.id,
        "name": course.name,
        "description": course.description,
        "difficulty": course.difficulty,
        "duration_minutes": course.duration_minutes,
        "scenario_ids": course.scenario_ids,
        "is_published": course.is_published,
        "created_at": course.created_at,
    }


@router.put("/courses/{course_id}")
async def update_course(
    course_id: str,
    course_data: CourseCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Update course"""
    course = db.query(Course).filter(Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    if course.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    course.name = course_data.name or course.name
    course.description = course_data.description or course.description
    course.scenario_ids = course_data.scenario_ids or course.scenario_ids
    course.updated_at = datetime.utcnow()
    db.commit()
    return {"status": "updated", "course_id": course_id}


@router.delete("/courses/{course_id}")
async def delete_course(
    course_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Delete course (only if not assigned to batches)"""
    course = db.query(Course).filter(Course.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    if course.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    assignments = db.query(CourseAssignment).filter(CourseAssignment.course_id == course_id).count()
    if assignments > 0:
        raise HTTPException(status_code=400, detail="Cannot delete course with active assignments")
    db.delete(course)
    db.commit()
    return {"status": "deleted", "course_id": course_id}


@router.post("/courses/{course_id}/publish")
async def publish_course(
    course_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Publish a course (make available for assignment)"""
    course = _get_trainer_course(db, trainer_id=current_user.id, course_id=course_id)

    course.is_published = True
    db.commit()

    return {"status": "published", "course_id": course_id}


# ==================== Course Assignment ====================


@router.post("/assign-course", response_model=dict)
async def assign_course(
    assignment: CourseAssignmentCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Assign a course to a batch or individual trainee"""
    course = _get_trainer_course(db, trainer_id=current_user.id, course_id=assignment.course_id)

    if bool(assignment.batch_id) == bool(assignment.user_id):
        raise HTTPException(
            status_code=400,
            detail="Specify either a batch_id or a user_id for the assignment.",
        )

    target_batch_id = None
    target_user_id = None
    if assignment.batch_id:
        batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=assignment.batch_id)
        target_batch_id = batch.id
    else:
        trainee = _get_trainer_trainee(db, trainer_id=current_user.id, trainee_id=assignment.user_id or "")
        target_user_id = trainee.id

    existing_assignment = (
        db.query(CourseAssignment)
        .filter(
            CourseAssignment.course_id == course.id,
            CourseAssignment.batch_id == target_batch_id,
            CourseAssignment.user_id == target_user_id,
            CourseAssignment.assigned_by == current_user.id,
        )
        .first()
    )
    if existing_assignment:
        return {
            "id": existing_assignment.id,
            "course_id": course.id,
            "assigned_to": target_batch_id or target_user_id,
            "status": "already_assigned",
        }

    new_assignment = CourseAssignment(
        course_id=course.id,
        batch_id=target_batch_id,
        user_id=target_user_id,
        assigned_by=current_user.id,
        is_mandatory=assignment.is_mandatory,
        due_date=assignment.due_date,
    )

    db.add(new_assignment)
    db.commit()

    return {
        "id": new_assignment.id,
        "course_id": course.id,
        "assigned_to": target_batch_id or target_user_id,
        "status": "assigned",
    }


@router.get("/course-assignments")
async def list_course_assignments(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
    batch_id: Optional[str] = None,
):
    """List course assignments for trainer's batches"""
    query = db.query(CourseAssignment).filter(CourseAssignment.assigned_by == current_user.id)

    if batch_id:
        batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)
        query = query.filter(CourseAssignment.batch_id == batch.id)

    assignments = query.all()

    return {
        "count": len(assignments),
        "assignments": [
            {
                "id": a.id,
                "course_id": a.course_id,
                "course_name": a.course.name if a.course else None,
                "batch_id": a.batch_id,
                "user_id": a.user_id,
                "assigned_at": a.assigned_at,
                "due_date": a.due_date,
                "is_mandatory": a.is_mandatory,
                "completion_percentage": a.completion_percentage,
            }
            for a in assignments
        ],
    }


# ==================== Interaction Review & Verification ====================


@router.get("/interaction-history")
async def get_interaction_history(
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
    batch_id: Optional[str] = None,
    user_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
):
    """Get practice sessions for trainer's assigned batch/users"""
    trainer_batches = db.query(Batch).filter(Batch.created_by == current_user.id).all()
    trainer_user_ids = {
        user.id
        for batch in trainer_batches
        for user in batch.users
        if user.role == UserRole.TRAINEE
    }

    if batch_id:
        batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)
        trainer_user_ids = {
            user.id for user in batch.users if user.role == UserRole.TRAINEE
        }

    if user_id:
        trainee = _get_trainer_trainee(db, trainer_id=current_user.id, trainee_id=user_id)
        trainer_user_ids = {trainee.id}

    if not trainer_user_ids:
        return {"count": 0, "sessions": []}

    query = db.query(PracticeSession).filter(PracticeSession.user_id.in_(trainer_user_ids))

    sessions = (
        query.order_by(PracticeSession.created_at.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )

    return {
        "count": len(sessions),
        "sessions": [
            {
                "id": s.id,
                "user_id": s.user_id,
                "user_name": s.user.full_name if s.user else None,
                "scenario_id": s.scenario_id,
                "scenario_title": s.scenario.title if s.scenario else None,
                "overall_score": s.overall_score,
                "accuracy": s.accuracy_score,
                "fluency": s.fluency_score,
                "attempt_number": s.attempt_number,
                "status": s.status,
                "is_verified": s.is_verified,
                "created_at": s.created_at,
                "audio_file_url": s.audio_file_url,
            }
            for s in sessions
        ],
    }


@router.get("/interactions/{session_id}")
async def get_interaction_detail(
    session_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Get detailed interaction information with transcript"""
    session = _get_trainer_session(db, trainer_id=current_user.id, session_id=session_id)

    return {
        "id": session.id,
        "user_id": session.user_id,
        "user_name": session.user.full_name if session.user else None,
        "scenario_id": session.scenario_id,
        "scenario_title": session.scenario.title if session.scenario else None,
        "audio_file_url": session.audio_file_url,
        "transcription": session.transcription,
        "transcription_confidence": session.transcription_confidence,
        "overall_score": session.overall_score,
        "scores": {
            "accuracy": session.accuracy_score,
            "fluency": session.fluency_score,
            "clarity": session.clarity_score,
            "keyword_adherence": session.keyword_adherence_score,
            "soft_skills": session.soft_skills_score,
        },
        "word_feedback": session.word_feedback,
        "filler_words": session.filler_words_detected,
        "assessment_data": session.assessment_data,
        "response_duration": session.response_duration,
        "dead_air_time": session.dead_air_time,
        "volume_level": session.volume_level,
        "attempt_number": session.attempt_number,
        "is_verified": session.is_verified,
        "feedback_items": [
            {
                "id": f.id,
                "feedback_type": f.feedback_type,
                "content": f.content,
                "recommended_module_id": f.recommended_module_id,
                "created_at": f.created_at,
            }
            for f in session.feedback_items
        ],
        "created_at": session.created_at,
    }


@router.post("/interactions/{session_id}/verify")
async def verify_interaction(
    session_id: str,
    verification: SessionVerification,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Verify/correct ASR transcription for a session"""
    session = _get_trainer_session(db, trainer_id=current_user.id, session_id=session_id)

    session.is_verified = verification.is_verified

    if verification.corrected_transcription:
        session.transcription = verification.corrected_transcription

    session.reviewed_by = current_user.id
    db.commit()

    return {"status": "verified", "session_id": session_id}


# ==================== Coaching & Feedback ====================


@router.post("/feedback", response_model=dict)
async def create_feedback(
    feedback_data: FeedbackCreate,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Provide coaching feedback on a practice session"""
    session = _get_trainer_session(
        db,
        trainer_id=current_user.id,
        session_id=feedback_data.practice_session_id,
    )

    new_feedback = Feedback(
        practice_session_id=session.id,
        trainer_id=current_user.id,
        feedback_type=feedback_data.feedback_type,
        content=feedback_data.content,
        recommended_module_id=feedback_data.recommended_module_id,
        recommended_exercises=feedback_data.recommended_exercises or [],
        is_automated=False,
    )

    db.add(new_feedback)
    db.commit()

    return {
        "id": new_feedback.id,
        "session_id": feedback_data.practice_session_id,
        "status": "created",
    }


@router.get("/feedback/{session_id}")
async def get_feedback(
    session_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Get all feedback for a session"""
    _get_trainer_session(db, trainer_id=current_user.id, session_id=session_id)
    feedback_items = (
        db.query(Feedback).filter(Feedback.practice_session_id == session_id).all()
    )

    return {
        "count": len(feedback_items),
        "feedback": [
            {
                "id": f.id,
                "feedback_type": f.feedback_type,
                "content": f.content,
                "trainer_name": f.trainer.full_name if f.trainer else None,
                "recommended_module_id": f.recommended_module_id,
                "is_acknowledged": f.is_acknowledge_by_trainee,
                "created_at": f.created_at,
            }
            for f in feedback_items
        ],
    }


@router.post("/feedback/{feedback_id}/push-module")
async def push_microlearning(
    feedback_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
):
    """Push a microlearning module to trainee based on feedback"""
    feedback = db.query(Feedback).filter(Feedback.id == feedback_id).first()

    if not feedback or not feedback.recommended_module_id:
        raise HTTPException(status_code=404, detail="Feedback or module not found")
    if feedback.trainer_id != current_user.id:
        raise HTTPException(status_code=404, detail="Feedback not found")

    _get_trainer_session(
        db,
        trainer_id=current_user.id,
        session_id=feedback.practice_session_id,
    )

    # Create assignment for trainee
    # This would be stored as a notification/assignment

    return {
        "status": "module_assigned",
        "module_id": feedback.recommended_module_id,
        "session_id": feedback.practice_session_id,
    }


# ==================== Performance Analysis ====================


@router.get("/batch-performance/{batch_id}")
async def get_batch_performance(
    batch_id: str, current_user: Any = Depends(verify_trainer), db: Session = Depends()
):
    """Get performance analytics for a batch"""
    batch = _get_trainer_batch(db, trainer_id=current_user.id, batch_id=batch_id)

    # Get all sessions for batch users
    user_ids = [u.id for u in batch.users if u.role == UserRole.TRAINEE]
    sessions = (
        db.query(PracticeSession).filter(PracticeSession.user_id.in_(user_ids)).all()
    )

    # Calculate metrics by user
    user_metrics = {}
    for session in sessions:
        if session.user_id not in user_metrics:
            user_metrics[session.user_id] = {
                "sessions": [],
                "avg_score": 0,
                "passed_count": 0,
                "failed_count": 0,
            }

        user_metrics[session.user_id]["sessions"].append(session)
        if session.overall_score and session.overall_score >= 70:
            user_metrics[session.user_id]["passed_count"] += 1
        else:
            user_metrics[session.user_id]["failed_count"] += 1

    # Calculate averages
    for metrics in user_metrics.values():
        if metrics["sessions"]:
            total_score = sum(s.overall_score or 0 for s in metrics["sessions"])
            metrics["avg_score"] = total_score / len(metrics["sessions"])

    return {
        "batch_id": batch_id,
        "batch_name": batch.name,
        "total_users": len(batch.users),
        "total_sessions": len(sessions),
        "user_performance": [
            {
                "user_id": user_id,
                "user_name": next(
                    (u.full_name for u in batch.users if u.id == user_id), None
                ),
                "avg_score": round(metrics["avg_score"], 2),
                "passed_sessions": metrics["passed_count"],
                "failed_sessions": metrics["failed_count"],
                "session_count": len(metrics["sessions"]),
            }
            for user_id, metrics in user_metrics.items()
        ],
    }


@router.get("/user-performance/{user_id}")
async def get_user_performance(
    user_id: str,
    current_user: Any = Depends(verify_trainer),
    db: Session = Depends(),
    days: int = 7,
):
    """Get detailed performance metrics for a trainee"""
    from datetime import timedelta

    from sqlalchemy import and_

    _get_trainer_trainee(db, trainer_id=current_user.id, trainee_id=user_id)

    # Get sessions in timeframe
    cutoff_date = datetime.utcnow() - timedelta(days=days)
    sessions = (
        db.query(PracticeSession)
        .filter(
            and_(
                PracticeSession.user_id == user_id,
                PracticeSession.created_at >= cutoff_date,
            )
        )
        .order_by(PracticeSession.created_at.asc())
        .all()
    )

    # Calculate metrics
    total_sessions = len(sessions)
    passed_sessions = sum((s.overall_score or 0) >= 70 for s in sessions)

    avg_scores = {
        "overall": (
            sum(s.overall_score or 0 for s in sessions) / total_sessions
            if total_sessions
            else 0
        ),
        "accuracy": (
            sum(s.accuracy_score or 0 for s in sessions) / total_sessions
            if total_sessions
            else 0
        ),
        "fluency": (
            sum(s.fluency_score or 0 for s in sessions) / total_sessions
            if total_sessions
            else 0
        ),
        "clarity": (
            sum(s.clarity_score or 0 for s in sessions) / total_sessions
            if total_sessions
            else 0
        ),
        "keyword_adherence": (
            sum(s.keyword_adherence_score or 0 for s in sessions) / total_sessions
            if total_sessions
            else 0
        ),
        "soft_skills": (
            sum(s.soft_skills_score or 0 for s in sessions) / total_sessions
            if total_sessions
            else 0
        ),
    }

    return {
        "user_id": user_id,
        "period_days": days,
        "total_sessions": total_sessions,
        "passed_sessions": passed_sessions,
        "failed_sessions": total_sessions - passed_sessions,
        "pass_rate": (
            round((passed_sessions / total_sessions * 100), 2)
            if total_sessions > 0
            else 0
        ),
        "average_scores": {k: round(v, 2) for k, v in avg_scores.items()},
        "session_timeline": [
            {
                "session_id": s.id,
                "scenario_title": s.scenario.title if s.scenario else None,
                "overall_score": s.overall_score,
                "created_at": s.created_at,
            }
            for s in sessions
        ],
    }


# ==================== Trainer Dashboard ====================


@router.websocket("/live-updates")
async def trainer_live_updates(websocket: WebSocket, token: str):
    """Push new trainee session events to trainer and admin dashboards."""
    db = SessionLocal()
    channel = "trainers"
    try:
        token_data = auth_utils.decode_token(token)
        user = db.query(User).filter(User.id == token_data.user_id).first()

        if not user or user.role not in (UserRole.TRAINER, UserRole.ADMIN):
            await websocket.close(code=4403)
            return

        if user.role == UserRole.ADMIN:
            channel = "admins"

        await live_update_manager.connect(channel, websocket)
        await websocket.send_json(
            {
                "type": "connected",
                "role": user.role.value,
                "message": "Live dashboard updates enabled",
            }
        )

        while True:
            message = await websocket.receive_text()
            if message == "ping":
                await websocket.send_json({"type": "pong"})
    except HTTPException:
        try:
            await websocket.accept()
        except Exception:
            pass
        try:
            await websocket.close(code=4401)
        except Exception:
            pass
    except WebSocketDisconnect:
        await live_update_manager.disconnect(channel, websocket)
    except Exception:
        try:
            await websocket.close(code=1011)
        except Exception:
            pass
        await live_update_manager.disconnect(channel, websocket)
    finally:
        db.close()


@router.get("/stats")
async def trainer_stats(
    current_user: Any = Depends(verify_trainer), db: Session = Depends()
):
    """Stats payload used by trainer dashboard UI."""
    data = await trainer_dashboard(current_user=current_user, db=db)
    return {
        "total_trainees": data.get("total_assigned_users", 0),
        "total_batches": data.get("total_batches", 0),
        "total_sessions": data.get("total_sessions", 0),
        "average_score": data.get("average_session_score", 0),
        "pending_reviews": data.get("pending_reviews", 0),
    }


@router.get("/dashboard")
async def trainer_dashboard(
    current_user: Any = Depends(verify_trainer), db: Session = Depends()
):
    """Trainer dashboard summary"""
    # Get batches
    batches = db.query(Batch).filter(Batch.created_by == current_user.id).all()
    batch_ids = [b.id for b in batches]

    # Get user IDs from batches
    from sqlalchemy import and_

    batch_user_ids = (
        db.query(User).filter(User.batches.any(Batch.id.in_(batch_ids))).all()
    )
    user_ids = [u.id for u in batch_user_ids]

    # Get sessions
    sessions = (
        db.query(PracticeSession).filter(PracticeSession.user_id.in_(user_ids)).all()
        if user_ids
        else []
    )

    # Calculate metrics
    total_users = len(batch_user_ids)
    total_sessions = len(sessions)

    avg_score = (
        sum(s.overall_score or 0 for s in sessions) / total_sessions if sessions else 0
    )

    return {
        "trainer_name": current_user.full_name,
        "total_batches": len(batches),
        "total_assigned_users": total_users,
        "total_sessions": total_sessions,
        "average_session_score": round(avg_score, 2),
        "pending_reviews": sum(not s.is_verified for s in sessions),
        "timestamp": datetime.utcnow(),
    }
